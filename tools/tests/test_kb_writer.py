"""Standalone tests for tools/kb_writer.py (UP-1.2 phase 2).

Run:
    python3 tools/tests/test_kb_writer.py

Each test copies the live KB to a temp dir + runs mutations against the copy.
The real KB is never touched. No external services required.
"""
from __future__ import annotations

import os
import shutil
import sys
import tempfile
import traceback
import unittest
from pathlib import Path

# Repo root in sys.path so `tools.*` imports resolve.
_HERE = Path(__file__).resolve().parent
_REPO_ROOT = _HERE.parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from openpyxl import load_workbook

from tools import kb_schema as S
from tools import kb_validators as V
from tools import kb_writer as W


class KBWriterTestBase(unittest.TestCase):
    """Copies the live KB into a temp dir per test for full isolation."""

    @classmethod
    def setUpClass(cls):
        cls.live_kb = S.DEFAULT_KB_PATH
        if not cls.live_kb.exists():
            raise unittest.SkipTest(f"Live KB not found at {cls.live_kb}")

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="kb_writer_test_"))
        self.kb_path = self.tmpdir / self.live_kb.name
        shutil.copy2(self.live_kb, self.kb_path)
        # Use temp backups dir to avoid polluting repo backups/
        self._orig_repo_root = S.REPO_ROOT
        self._backups_dir = self.tmpdir / "backups"
        self._backups_dir.mkdir(exist_ok=True)
        S.REPO_ROOT = self.tmpdir  # type: ignore

        # Normalize preexisting KB quirks that would fail validate_all but are
        # orthogonal to UP-1.2 phase 2 scope. Cases row 9 has case_quality
        # "SILVER (LIVE - outcome pending end of day)" which validate_case_quality
        # rejects. Replace with bare "SILVER" for test isolation.
        # (The real KB still has the bad value — separate cleanup PR.)
        wb_norm = load_workbook(self.kb_path, data_only=False, read_only=False)
        if S.CASES_SHEET in wb_norm.sheetnames:
            ws_cases = wb_norm[S.CASES_SHEET]
            # Find case_quality col.
            cq_col = None
            for col_idx, cell in enumerate(ws_cases[S.CASES_HEADER_ROW], start=1):
                if cell.value and str(cell.value).strip() == "case_quality":
                    cq_col = col_idx
                    break
            if cq_col is not None:
                for row_idx in range(S.CASES_DATA_START_ROW, ws_cases.max_row + 1):
                    cell = ws_cases.cell(row=row_idx, column=cq_col)
                    if cell.value is None:
                        continue
                    val = str(cell.value).strip().upper()
                    if val not in {"GOLD", "SILVER", "BRONZE"}:
                        # Coerce to bare GOLD/SILVER/BRONZE if it starts with one.
                        for valid in ("GOLD", "SILVER", "BRONZE"):
                            if val.startswith(valid):
                                cell.value = valid
                                break
        wb_norm.save(self.kb_path)

    def tearDown(self):
        S.REPO_ROOT = self._orig_repo_root  # type: ignore
        shutil.rmtree(self.tmpdir, ignore_errors=True)


class TestAddRule(KBWriterTestBase):
    def test_add_rule_appends_and_validates(self):
        backup = W.add_rule(
            kb_path=self.kb_path,
            rule_id="TR-Juan-999",
            tier="TACTICAL_PLUS",
            trigger="Test trigger condition",
            action="Test action: SELL_PUT delta 0.10",
            confidence="MEDIUM",
            notes="UP-1.2 test add",
        )
        self.assertTrue(backup.exists())

        # Verify row written.
        wb = V.load_kb(self.kb_path)
        _, row = W.find_rule_row(wb, "TR-Juan-999")
        self.assertEqual(V.extract_rule_id(row[0]), "TR-Juan-999")
        tier_col = S.DECISION_RULES_COLS["tier"] - 1
        self.assertEqual(str(row[tier_col]).upper(), "TACTICAL_PLUS")

        # Validate KB still passes.
        report = V.validate_all(wb)
        for check, errs in report.items():
            self.assertEqual(errs, [], f"{check} errors: {errs}")

    def test_add_rule_rejects_invalid_tier(self):
        with self.assertRaises(ValueError):
            W.add_rule(
                kb_path=self.kb_path,
                rule_id="TR-Juan-998",
                tier="INVALID_TIER",
                trigger="t",
                action="a",
                confidence="HIGH",
            )

    def test_add_rule_rejects_duplicate_id(self):
        # TR-Juan-001 already exists in the KB.
        with self.assertRaises((ValueError, W.KBValidationError)):
            W.add_rule(
                kb_path=self.kb_path,
                rule_id="TR-Juan-001",
                tier="TACTICAL",
                trigger="dup",
                action="dup",
                confidence="LOW",
            )


class TestEditRule(KBWriterTestBase):
    def test_edit_rule_updates_field(self):
        backup, before, after = W.edit_rule(
            kb_path=self.kb_path,
            rule_id="TR-Juan-001",
            field="Notes",
            value="UP-1.2 test edit",
        )
        self.assertTrue(backup.exists())
        self.assertEqual(after, "UP-1.2 test edit")

        wb = V.load_kb(self.kb_path)
        _, row = W.find_rule_row(wb, "TR-Juan-001")
        notes_col = S.DECISION_RULES_COLS["Notes"] - 1
        self.assertEqual(str(row[notes_col]), "UP-1.2 test edit")

    def test_edit_rule_rejects_rule_id_field(self):
        with self.assertRaises(ValueError):
            W.edit_rule(
                kb_path=self.kb_path,
                rule_id="TR-Juan-001",
                field="Rule_ID",
                value="TR-Juan-001-renamed",
            )


class TestDeleteRule(KBWriterTestBase):
    def test_delete_rule_soft_marks_deprecated(self):
        backup = W.delete_rule_soft(
            kb_path=self.kb_path,
            rule_id="TR-Juan-001",
            reason="test soft delete",
        )
        self.assertTrue(backup.exists())

        wb = V.load_kb(self.kb_path)
        _, row = W.find_rule_row(wb, "TR-Juan-001")  # still findable
        notes_col = S.DECISION_RULES_COLS["Notes"] - 1
        validated_col = S.DECISION_RULES_COLS["Validated?"] - 1
        notes = str(row[notes_col])
        self.assertIn("DEPRECATED", notes)
        self.assertIn("test soft delete", notes)
        self.assertEqual(str(row[validated_col]).lower(), "false")

    def test_delete_rule_hard_removes_row(self):
        # Choose a rule unlikely to have ghost refs: highest id in the KB.
        wb_before = V.load_kb(self.kb_path)
        next_id = W.compute_next_rule_id(wb_before)
        # Add a synthetic rule we can safely hard-delete.
        W.add_rule(
            kb_path=self.kb_path,
            rule_id=next_id,
            tier="TACTICAL",
            trigger="ephemeral",
            action="ephemeral",
            confidence="LOW",
            notes="for hard-delete test",
        )

        backup, ghosts = W.delete_rule_hard(
            kb_path=self.kb_path,
            rule_id=next_id,
            reason="ephemeral test",
        )
        self.assertTrue(backup.exists())
        self.assertEqual(ghosts, [])

        wb = V.load_kb(self.kb_path)
        with self.assertRaises(LookupError):
            W.find_rule_row(wb, next_id)

    def test_delete_rule_hard_requires_reason(self):
        with self.assertRaises(ValueError):
            W.delete_rule_hard(self.kb_path, "TR-Juan-001", reason="")


class TestMergeRules(KBWriterTestBase):
    def test_merge_rules_combines_and_marks_source_deprecated(self):
        wb_before = V.load_kb(self.kb_path)
        # Use 2 synthetic rules to avoid affecting real KB content.
        a_id = W.compute_next_rule_id(wb_before)
        W.add_rule(self.kb_path, a_id, "TACTICAL", "trigA", "actA", "LOW", notes="A")
        wb_mid = V.load_kb(self.kb_path)
        b_id = W.compute_next_rule_id(wb_mid)
        W.add_rule(self.kb_path, b_id, "TACTICAL", "trigB", "actB", "LOW", notes="B")

        # Merge B into A (lower id by default).
        backup = W.merge_rules(
            kb_path=self.kb_path,
            rule_ids=[a_id, b_id],
            into=a_id,
            reason="test merge",
        )
        self.assertTrue(backup.exists())

        wb = V.load_kb(self.kb_path)
        _, row_a = W.find_rule_row(wb, a_id)
        _, row_b = W.find_rule_row(wb, b_id)

        trig_col = S.DECISION_RULES_COLS["Trigger Conditions"] - 1
        act_col = S.DECISION_RULES_COLS["Action"] - 1
        notes_col = S.DECISION_RULES_COLS["Notes"] - 1
        val_col = S.DECISION_RULES_COLS["Validated?"] - 1

        # Target has concatenated triggers + actions.
        self.assertIn("trigA", str(row_a[trig_col]))
        self.assertIn("trigB", str(row_a[trig_col]))
        self.assertIn("actA", str(row_a[act_col]))
        self.assertIn("actB", str(row_a[act_col]))
        self.assertIn("MERGED", str(row_a[notes_col]))

        # Source is deprecated.
        self.assertIn("DEPRECATED", str(row_b[notes_col]))
        self.assertEqual(str(row_b[val_col]).lower(), "false")


class TestBumpVersion(KBWriterTestBase):
    @staticmethod
    def _versions_from_kb_name(kb_name: str) -> tuple[str, str]:
        """Derive (from_version, to_version) from a KB filename like
        EOLO_ThetaHarvest_v1.3.xlsx → ('v1.3', 'v1.4').

        Keeps the bump tests version-agnostic so they survive future bumps
        without manual edits.
        """
        import re
        m = re.search(r"v(\d+)\.(\d+)", kb_name)
        if not m:
            raise ValueError(f"Cannot extract version from {kb_name}")
        major, minor = int(m.group(1)), int(m.group(2))
        return f"v{major}.{minor}", f"v{major}.{minor + 1}"

    def test_bump_version_renames_and_creates_backup(self):
        from_v, to_v = self._versions_from_kb_name(self.kb_path.name)
        # Build a self-contained test: copy the kb_loader.py and kb_schema.py
        # into the temp REPO_ROOT so _replace_in_file can act on them.
        eng_dir = self.tmpdir / "llm_engine_eolo" / "llm_engine"
        eng_dir.mkdir(parents=True, exist_ok=True)
        (eng_dir / "kb_loader.py").write_text(
            f'KB_VERSION = "{from_v}"\n', encoding="utf-8"
        )
        tools_dir = self.tmpdir / "tools"
        tools_dir.mkdir(exist_ok=True)
        (tools_dir / "kb_schema.py").write_text(
            f'DEFAULT_KB_PATH = "EOLO_ThetaHarvest_{from_v}.xlsx"\n', encoding="utf-8"
        )

        backup, new_path = W.bump_version(
            kb_path=self.kb_path,
            from_version=from_v,
            to_version=to_v,
        )
        self.assertTrue(backup.exists())
        self.assertTrue(new_path.exists())
        self.assertIn(to_v, new_path.name)
        self.assertFalse(self.kb_path.exists())  # old name gone

        # Refs updated in stub files.
        loader_content = (eng_dir / "kb_loader.py").read_text(encoding="utf-8")
        self.assertIn(to_v, loader_content)
        self.assertNotIn(from_v, loader_content)

    def test_bump_version_skips_code_refs_when_disabled(self):
        from_v, to_v = self._versions_from_kb_name(self.kb_path.name)
        # Stub files that should NOT be modified when update_code_refs=False.
        eng_dir = self.tmpdir / "llm_engine_eolo" / "llm_engine"
        eng_dir.mkdir(parents=True, exist_ok=True)
        loader_path = eng_dir / "kb_loader.py"
        loader_path.write_text(f'KB_VERSION = "{from_v}"\n', encoding="utf-8")
        tools_dir = self.tmpdir / "tools"
        tools_dir.mkdir(exist_ok=True)
        schema_path = tools_dir / "kb_schema.py"
        schema_path.write_text(
            f'DEFAULT_KB_PATH = "EOLO_ThetaHarvest_{from_v}.xlsx"\n', encoding="utf-8"
        )

        backup, new_path = W.bump_version(
            kb_path=self.kb_path,
            from_version=from_v,
            to_version=to_v,
            update_code_refs=False,
        )

        # KB rename + backup still happen.
        self.assertTrue(backup.exists())
        self.assertTrue(new_path.exists())
        self.assertIn(to_v, new_path.name)
        self.assertFalse(self.kb_path.exists())

        # But code refs are untouched.
        self.assertEqual(
            loader_path.read_text(encoding="utf-8"), f'KB_VERSION = "{from_v}"\n'
        )
        self.assertEqual(
            schema_path.read_text(encoding="utf-8"),
            f'DEFAULT_KB_PATH = "EOLO_ThetaHarvest_{from_v}.xlsx"\n',
        )


class TestBackupAtomic(KBWriterTestBase):
    def test_backup_creates_valid_file(self):
        backup = W.backup_kb(self.kb_path, "test_suffix")
        self.assertTrue(backup.exists())
        self.assertGreater(backup.stat().st_size, 0)
        # Backup is a valid KB.
        wb = V.load_kb(backup)
        report = V.validate_all(wb)
        for check, errs in report.items():
            self.assertEqual(errs, [], f"backup invalid for {check}: {errs}")


class TestValidationRollback(KBWriterTestBase):
    def test_add_rule_invalid_post_write_restores_backup(self):
        # Force a post-write validation failure by directly corrupting the KB
        # inside the mutation. We do this via with_backup_validate with a
        # malicious mutate that adds an unknown tier directly.
        def _bad_mutate(wb):
            # Append a row with invalid tier (bypasses add_rule's pre-check).
            W.append_row(
                wb,
                S.DECISION_RULES_SHEET,
                {
                    "Rule_ID": "TR-Juan-997",
                    "Trigger Conditions": "x",
                    "Action": "x",
                    "Confidence Required": "HIGH",
                    "Source Cases": "",
                    "Validated?": "false",
                    "Notes": "",
                    "tier": "BOGUS_TIER",
                },
                header_row=S.DECISION_RULES_HEADER_ROW,
            )

        original_size = self.kb_path.stat().st_size
        with self.assertRaises(W.KBValidationError):
            W.with_backup_validate(self.kb_path, "test_invalid", _bad_mutate)
        # KB restored from backup.
        self.assertTrue(self.kb_path.exists())
        # Confirm rule was NOT persisted (rollback worked).
        wb = V.load_kb(self.kb_path)
        with self.assertRaises(LookupError):
            W.find_rule_row(wb, "TR-Juan-997")


def _run() -> int:
    """Run all tests + return process exit code."""
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    for cls in (
        TestAddRule,
        TestEditRule,
        TestDeleteRule,
        TestMergeRules,
        TestBumpVersion,
        TestBackupAtomic,
        TestValidationRollback,
    ):
        suite.addTests(loader.loadTestsFromTestCase(cls))
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    return 0 if result.wasSuccessful() else 1


if __name__ == "__main__":
    sys.exit(_run())
