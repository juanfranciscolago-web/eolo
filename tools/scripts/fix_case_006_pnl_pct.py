"""One-off fix #93: Case #006 pnl_pct normalize.

Case 2026-05-27_SPY_counterfactual_006 quedó con pnl_pct = "PENDING_VALIDATION_TODAY..."
porque fue counterfactual (trade_executed='NO', sin fill real en eolo-crop-trades).
Decisión 2026-06-01: marcar pnl_pct como N/A definitivo + append a notes con
justificación. Valor del case está en lesson_learned (cualitativo), no en pnl.

Run desde repo root:
    python3 tools/scripts/fix_case_006_pnl_pct.py
"""
from pathlib import Path
import sys

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

from tools import kb_writer as W  # noqa: E402

KB_PATH = REPO_ROOT / "llm_engine_eolo" / "kb" / "EOLO_ThetaHarvest_v1.2.xlsx"
CASE_ID = "2026-05-27_SPY_counterfactual_006"

OLD_PNL = "PENDING_VALIDATION_TODAY - revisar a las 16:00 ET 27-may"
NEW_PNL = "N/A (counterfactual sin simulación)"

NOTES_APPEND = (
    " [SPRINT #93 cleanup 2026-06-01: pnl_pct marcado N/A definitivo. "
    "Case counterfactual sin trade real (trade_executed=NO); valor cualitativo "
    "en lesson_learned. Sin simulación BS para mantener honestidad.]"
)


def main() -> int:
    if not KB_PATH.exists():
        print(f"ERROR: KB not found at {KB_PATH}", file=sys.stderr)
        return 1

    backup = W.backup_kb(KB_PATH, "pre_case_006_pnl_fix")
    print(f"Backup: {backup}")

    wb = openpyxl.load_workbook(KB_PATH)
    ws = wb["Cases"]

    # Locate row by case_id (col 1).
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == CASE_ID:
            target_row = row_idx
            break
    if target_row is None:
        print(f"ERROR: case_id '{CASE_ID}' not found in Cases sheet", file=sys.stderr)
        return 1
    print(f"Found case at row {target_row}")

    # Locate pnl_pct cell (value == OLD_PNL) + notes cell (starts with 'CASO EN VIVO').
    pnl_col = None
    notes_col = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=target_row, column=col_idx).value
        if val == OLD_PNL:
            pnl_col = col_idx
        elif isinstance(val, str) and val.startswith("CASO EN VIVO"):
            notes_col = col_idx
    if pnl_col is None:
        print(f"ERROR: pnl_pct cell with old value not found in row {target_row}", file=sys.stderr)
        return 1
    if notes_col is None:
        print(f"ERROR: notes cell starting with 'CASO EN VIVO' not found in row {target_row}", file=sys.stderr)
        return 1
    print(f"  pnl_pct cell: column {pnl_col}")
    print(f"  notes   cell: column {notes_col}")

    # Write.
    ws.cell(row=target_row, column=pnl_col).value = NEW_PNL
    old_notes = ws.cell(row=target_row, column=notes_col).value
    ws.cell(row=target_row, column=notes_col).value = old_notes + NOTES_APPEND

    W.save_workbook_atomic(wb, KB_PATH)
    print(f"✅ Updated Case #006 pnl_pct in {KB_PATH}")
    print(f"   Old pnl_pct: {OLD_PNL!r}")
    print(f"   New pnl_pct: {NEW_PNL!r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
