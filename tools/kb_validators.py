"""KB schema and integrity validators.

Each validator returns a list of human-readable error messages. An empty list
means the check passed. `validate_all` aggregates every check into a dict
keyed by check name so the CLI can present a structured report.
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from tools import kb_schema as S


# ── Loading ───────────────────────────────────────────────────────────────
def load_kb(path: str | Path) -> Workbook:
    """Open the KB workbook in data-only mode.

    Args:
        path: Filesystem path to the .xlsx file.

    Returns:
        A loaded openpyxl Workbook with formulas resolved to values.

    Raises:
        FileNotFoundError: If the file does not exist at `path`.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"KB Excel not found at {p}")
    return load_workbook(p, data_only=True, read_only=False)


def extract_rule_id(raw: object) -> str | None:
    """Normalise a raw Rule_ID cell to its canonical 'TR-Juan-NNN' form.

    The workbook stores values like 'TR-Juan-048 ⭐' or
    'TR-Juan-049 ⭐⭐ PROHIBITIVA'. We strip everything except the canonical id.

    Args:
        raw: The cell value (any type — coerced to str).

    Returns:
        The normalised id (e.g. 'TR-Juan-048') or None if no match.
    """
    if raw is None:
        return None
    m = S.RULE_ID_REGEX.search(str(raw))
    if not m:
        return None
    return S.RULE_ID_FORMAT.format(num=int(m.group(1)))


# ── Sheet-level validators ────────────────────────────────────────────────
def validate_sheets(wb: Workbook) -> list[str]:
    """Verify every expected sheet exists in the workbook.

    Args:
        wb: Loaded workbook.

    Returns:
        List of error strings (one per missing sheet); empty if all present.
    """
    missing = [s for s in S.EXPECTED_SHEETS if s not in wb.sheetnames]
    return [f"Missing sheet: '{name}'" for name in missing]


def validate_decision_rules_schema(wb: Workbook) -> list[str]:
    """Verify the Decision_Rules header row matches the expected layout.

    Args:
        wb: Loaded workbook.

    Returns:
        List of header-mismatch errors; empty if exact.
    """
    if S.DECISION_RULES_SHEET not in wb.sheetnames:
        return [f"Sheet '{S.DECISION_RULES_SHEET}' missing"]

    ws = wb[S.DECISION_RULES_SHEET]
    actual = [
        (str(c.value).strip() if c.value is not None else "")
        for c in ws[S.DECISION_RULES_HEADER_ROW]
    ]
    # Trim trailing empties so a wider sheet doesn't fail spuriously.
    while actual and actual[-1] == "":
        actual.pop()

    errors: list[str] = []
    if actual != S.DECISION_RULES_EXPECTED_HEADERS:
        errors.append(
            f"Decision_Rules header mismatch.\n"
            f"  expected: {S.DECISION_RULES_EXPECTED_HEADERS}\n"
            f"  actual:   {actual}"
        )
    return errors


def validate_unique_rule_ids(wb: Workbook) -> list[str]:
    """Verify every Rule_ID in Decision_Rules is canonical and unique.

    Args:
        wb: Loaded workbook.

    Returns:
        Errors for unparseable ids and duplicates.
    """
    if S.DECISION_RULES_SHEET not in wb.sheetnames:
        return [f"Sheet '{S.DECISION_RULES_SHEET}' missing"]

    ws = wb[S.DECISION_RULES_SHEET]
    seen: Counter[str] = Counter()
    errors: list[str] = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=S.DECISION_RULES_DATA_START_ROW, values_only=True),
        start=S.DECISION_RULES_DATA_START_ROW,
    ):
        if not row or row[0] is None:
            continue
        raw = str(row[0]).strip()
        if not raw.startswith("TR-"):
            continue
        rid = extract_rule_id(raw)
        if rid is None:
            errors.append(f"Row {row_idx}: cannot parse Rule_ID from '{raw}'")
            continue
        seen[rid] += 1

    for rid, count in seen.items():
        if count > 1:
            errors.append(f"Duplicate Rule_ID: {rid} (appears {count}x)")
    return errors


def validate_tiers(wb: Workbook) -> list[str]:
    """Verify the `tier` column is populated with a known tier on every rule.

    Args:
        wb: Loaded workbook.

    Returns:
        Errors for rows missing a tier or carrying an unknown one.
    """
    if S.DECISION_RULES_SHEET not in wb.sheetnames:
        return [f"Sheet '{S.DECISION_RULES_SHEET}' missing"]

    ws = wb[S.DECISION_RULES_SHEET]
    tier_col_idx = S.DECISION_RULES_COLS["tier"] - 1  # 0-indexed for tuple access
    errors: list[str] = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=S.DECISION_RULES_DATA_START_ROW, values_only=True),
        start=S.DECISION_RULES_DATA_START_ROW,
    ):
        if not row or row[0] is None:
            continue
        raw = str(row[0]).strip()
        if not raw.startswith("TR-"):
            continue
        rid = extract_rule_id(raw) or raw
        if len(row) <= tier_col_idx or row[tier_col_idx] is None:
            errors.append(f"{rid} (row {row_idx}): missing tier")
            continue
        tier = str(row[tier_col_idx]).strip().upper()
        if tier not in S.VALID_TIERS:
            errors.append(
                f"{rid} (row {row_idx}): unknown tier '{tier}' "
                f"(valid: {sorted(S.VALID_TIERS)})"
            )
    return errors


def validate_cases_schema(wb: Workbook) -> list[str]:
    """Verify the Cases header row contains every required field.

    Args:
        wb: Loaded workbook.

    Returns:
        Errors listing any required header that is missing.
    """
    if S.CASES_SHEET not in wb.sheetnames:
        return [f"Sheet '{S.CASES_SHEET}' missing"]

    ws = wb[S.CASES_SHEET]
    actual = {
        (str(c.value).strip() if c.value is not None else "")
        for c in ws[S.CASES_HEADER_ROW]
    }
    missing = sorted(S.CASES_REQUIRED_HEADERS - actual)
    return [f"Cases sheet missing required column: '{name}'" for name in missing]


def _cases_header_index(wb: Workbook) -> dict[str, int]:
    """Build a {header_name: 1-indexed column} map for the Cases sheet."""
    ws = wb[S.CASES_SHEET]
    out: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[S.CASES_HEADER_ROW], start=1):
        if cell.value:
            out[str(cell.value).strip()] = col_idx
    return out


def validate_case_quality(wb: Workbook) -> list[str]:
    """Verify every case row uses a known case_quality value.

    Args:
        wb: Loaded workbook.

    Returns:
        Errors for case rows with empty or unknown case_quality.
    """
    if S.CASES_SHEET not in wb.sheetnames:
        return [f"Sheet '{S.CASES_SHEET}' missing"]

    headers = _cases_header_index(wb)
    if "case_quality" not in headers or "case_id" not in headers:
        return ["Cases: case_quality or case_id column missing — schema check first"]

    ws = wb[S.CASES_SHEET]
    cq_idx = headers["case_quality"] - 1
    cid_idx = headers["case_id"] - 1
    errors: list[str] = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=S.CASES_DATA_START_ROW, values_only=True),
        start=S.CASES_DATA_START_ROW,
    ):
        if not row or cid_idx >= len(row) or row[cid_idx] is None:
            continue
        case_id = str(row[cid_idx]).strip()
        if cq_idx >= len(row) or row[cq_idx] is None:
            errors.append(f"{case_id} (row {row_idx}): missing case_quality")
            continue
        quality = str(row[cq_idx]).strip().upper()
        if quality not in S.VALID_CASE_QUALITY:
            errors.append(
                f"{case_id} (row {row_idx}): unknown case_quality '{quality}' "
                f"(valid: {sorted(S.VALID_CASE_QUALITY)})"
            )
    return errors


def validate_rule_references(wb: Workbook) -> list[str]:
    """Verify every TR-Juan-NNN cited in Cases.tacit_rules_applied exists.

    Args:
        wb: Loaded workbook.

    Returns:
        Errors listing ghost references per case.
    """
    if S.DECISION_RULES_SHEET not in wb.sheetnames or S.CASES_SHEET not in wb.sheetnames:
        return []

    defined: set[str] = set()
    for row in wb[S.DECISION_RULES_SHEET].iter_rows(
        min_row=S.DECISION_RULES_DATA_START_ROW, values_only=True
    ):
        if not row or row[0] is None:
            continue
        rid = extract_rule_id(row[0])
        if rid:
            defined.add(rid)

    headers = _cases_header_index(wb)
    if "tacit_rules_applied" not in headers or "case_id" not in headers:
        return []

    tra_idx = headers["tacit_rules_applied"] - 1
    cid_idx = headers["case_id"] - 1
    errors: list[str] = []

    for row in wb[S.CASES_SHEET].iter_rows(
        min_row=S.CASES_DATA_START_ROW, values_only=True
    ):
        if not row or cid_idx >= len(row) or row[cid_idx] is None:
            continue
        case_id = str(row[cid_idx]).strip()
        if tra_idx >= len(row) or row[tra_idx] is None:
            continue
        cited = {
            S.RULE_ID_FORMAT.format(num=int(m))
            for m in S.RULE_ID_REGEX.findall(str(row[tra_idx]))
        }
        ghosts = sorted(cited - defined)
        if ghosts:
            errors.append(f"Case {case_id}: ghost rule references {ghosts}")
    return errors


# ── Aggregator ────────────────────────────────────────────────────────────
def validate_all(wb: Workbook) -> dict[str, list[str]]:
    """Run every validator and return a {check_name: errors} report.

    Args:
        wb: Loaded workbook.

    Returns:
        Dict mapping each validator's short name to its error list.
    """
    return {
        "sheets": validate_sheets(wb),
        "decision_rules_schema": validate_decision_rules_schema(wb),
        "unique_rule_ids": validate_unique_rule_ids(wb),
        "tiers": validate_tiers(wb),
        "cases_schema": validate_cases_schema(wb),
        "case_quality": validate_case_quality(wb),
        "rule_references": validate_rule_references(wb),
    }
