"""KB write helpers (UP-1.2 phase 2).

Atomic mutations to the EOLO ThetaHarvest KB Excel. Every mutation:
  1. Creates a timestamped backup in `backups/` before writing.
  2. Saves via tempfile + os.replace (atomic on same filesystem).
  3. Re-validates the resulting workbook; restores backup if any check fails.

This module is read-only-by-default — nothing here is called from the bot
runtime. Used by tools/kb_editor.py CLI commands.
"""
from __future__ import annotations

import os
import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from tools import kb_schema as S
from tools import kb_validators as V


# ── Backup ────────────────────────────────────────────────────────────────
def backup_kb(kb_path: Path, suffix: Optional[str] = None) -> Path:
    """Copy `kb_path` to backups/ with a timestamp + optional suffix.

    Args:
        kb_path: Source KB workbook.
        suffix: Optional short tag to append (e.g. "pre_add_TR-062").

    Returns:
        Absolute path to the created backup file.

    Raises:
        FileNotFoundError: If `kb_path` does not exist.
    """
    if not kb_path.exists():
        raise FileNotFoundError(f"KB not found: {kb_path}")

    backups_dir = S.REPO_ROOT / "backups"
    backups_dir.mkdir(exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    parts = [kb_path.stem]
    if suffix:
        parts.append(suffix.replace(" ", "_"))
    parts.append(ts)
    backup_name = "_".join(parts) + kb_path.suffix
    backup_path = backups_dir / backup_name

    shutil.copy2(kb_path, backup_path)
    return backup_path


# ── Atomic save ───────────────────────────────────────────────────────────
def save_workbook_atomic(wb: Workbook, target_path: Path) -> None:
    """Save `wb` to `target_path` atomically via tempfile + os.replace.

    Args:
        wb: Workbook to save.
        target_path: Final destination .xlsx file.
    """
    target_path = Path(target_path)
    target_dir = target_path.parent
    target_dir.mkdir(parents=True, exist_ok=True)

    # tempfile en el mismo dir → garantiza same filesystem → os.replace atomic.
    fd, tmp_str = tempfile.mkstemp(
        prefix=f".{target_path.stem}_",
        suffix=target_path.suffix,
        dir=target_dir,
    )
    os.close(fd)
    tmp_path = Path(tmp_str)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, target_path)
    except Exception:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)
        raise


# ── Row operations ────────────────────────────────────────────────────────
def _header_to_col_index(ws, header_row: int, header_name: str) -> int:
    """Return 1-indexed column for `header_name` in `header_row`.

    Raises:
        KeyError: If header not found.
    """
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value is not None and str(cell.value).strip() == header_name:
            return col_idx
    raise KeyError(f"Header '{header_name}' not found in row {header_row}")


def update_row(
    wb: Workbook,
    sheet_name: str,
    row_index: int,
    field_updates: dict[str, Any],
    header_row: int = 1,
) -> None:
    """Update specific cells in `row_index` keyed by header name.

    Args:
        wb: Loaded workbook (mutated in place).
        sheet_name: Target sheet name.
        row_index: 1-indexed row to update.
        field_updates: Mapping of header name -> new value.
        header_row: Row that contains header names (default 1).
    """
    ws = wb[sheet_name]
    for header, value in field_updates.items():
        col = _header_to_col_index(ws, header_row, header)
        ws.cell(row=row_index, column=col, value=value)


def append_row(
    wb: Workbook,
    sheet_name: str,
    row_data: dict[str, Any],
    header_row: int = 1,
) -> int:
    """Append a new row to `sheet_name` keyed by header name.

    Args:
        wb: Loaded workbook (mutated).
        sheet_name: Target sheet name.
        row_data: Mapping of header name -> value for the new row.
        header_row: Row that contains header names.

    Returns:
        1-indexed row number where the new row was written.
    """
    ws = wb[sheet_name]
    # Encontrar primera fila vacía después del header.
    new_row = ws.max_row + 1
    # Si max_row == header_row (sheet vacío), comenzamos justo abajo.
    if new_row <= header_row:
        new_row = header_row + 1
    for header, value in row_data.items():
        col = _header_to_col_index(ws, header_row, header)
        ws.cell(row=new_row, column=col, value=value)
    return new_row


def remove_row(wb: Workbook, sheet_name: str, row_index: int) -> None:
    """Physically remove `row_index` from `sheet_name` (shifts rows up).

    Args:
        wb: Loaded workbook (mutated).
        sheet_name: Target sheet name.
        row_index: 1-indexed row to remove.
    """
    ws = wb[sheet_name]
    ws.delete_rows(row_index, amount=1)


# ── Rule_ID helpers ───────────────────────────────────────────────────────
def compute_next_rule_id(wb: Workbook) -> str:
    """Return the next free TR-Juan-NNN (max + 1 across Decision_Rules).

    Args:
        wb: Loaded workbook.

    Returns:
        Canonical id string, e.g. "TR-Juan-062".
    """
    max_num = 0
    if S.DECISION_RULES_SHEET not in wb.sheetnames:
        return S.RULE_ID_FORMAT.format(num=1)
    ws = wb[S.DECISION_RULES_SHEET]
    for row in ws.iter_rows(
        min_row=S.DECISION_RULES_DATA_START_ROW, values_only=True
    ):
        if not row or row[0] is None:
            continue
        m = S.RULE_ID_REGEX.search(str(row[0]))
        if m:
            max_num = max(max_num, int(m.group(1)))
    return S.RULE_ID_FORMAT.format(num=max_num + 1)


def find_rule_row(wb: Workbook, rule_id: str) -> tuple[int, tuple]:
    """Locate a rule by canonical id. Returns (row_index, row_tuple).

    Raises:
        LookupError: If rule_id not found.
    """
    target = V.extract_rule_id(rule_id)
    if target is None:
        raise LookupError(f"'{rule_id}' is not a valid TR-Juan-NNN")

    if S.DECISION_RULES_SHEET not in wb.sheetnames:
        raise LookupError(f"Sheet '{S.DECISION_RULES_SHEET}' missing")

    ws = wb[S.DECISION_RULES_SHEET]
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=S.DECISION_RULES_DATA_START_ROW, values_only=True),
        start=S.DECISION_RULES_DATA_START_ROW,
    ):
        if not row or row[0] is None:
            continue
        rid = V.extract_rule_id(row[0])
        if rid == target:
            return row_idx, row
    raise LookupError(f"Rule {target} not found in Decision_Rules")


# ── Validation harness ────────────────────────────────────────────────────
class KBValidationError(RuntimeError):
    """Raised when post-write validation fails."""


def with_backup_validate(
    kb_path: Path,
    suffix: str,
    mutate: Callable[[Workbook], None],
) -> Path:
    """Run a mutation with backup + atomic save + post-validate + auto-restore.

    Flow:
        1. backup_kb(kb_path, suffix) -> backup_path
        2. Load wb data_only=False (preserve formulas)
        3. mutate(wb)
        4. save_workbook_atomic(wb, kb_path)
        5. Re-load + validate_all
        6. If errors -> restore backup, raise KBValidationError
        7. Return backup_path on success

    Args:
        kb_path: Target KB file.
        suffix: Backup tag (e.g. "pre_add_TR-062").
        mutate: Callable that takes the workbook and mutates it.

    Returns:
        Path to the (preserved) backup file.

    Raises:
        KBValidationError: If validation fails after mutation.
        Any exception from `mutate` (with auto-restore from backup).
    """
    backup_path = backup_kb(kb_path, suffix)
    try:
        wb = load_workbook(kb_path, data_only=False, read_only=False)
        mutate(wb)
        save_workbook_atomic(wb, kb_path)
    except Exception:
        # mutate o save fallaron — restore backup.
        shutil.copy2(backup_path, kb_path)
        raise

    # Re-validate post-write.
    try:
        wb_check = V.load_kb(kb_path)
        report = V.validate_all(wb_check)
        errors = {check: errs for check, errs in report.items() if errs}
        if errors:
            shutil.copy2(backup_path, kb_path)
            raise KBValidationError(
                f"Post-write validation failed; backup restored from {backup_path}. "
                f"Errors: {errors}"
            )
    except KBValidationError:
        raise
    except Exception as e:
        # Failed to validate at all — restore as defensive measure.
        shutil.copy2(backup_path, kb_path)
        raise KBValidationError(
            f"Could not run post-write validation; backup restored. Reason: {e}"
        )

    return backup_path


# ── High-level mutations ──────────────────────────────────────────────────
def add_rule(
    kb_path: Path,
    rule_id: str,
    tier: str,
    trigger: str,
    action: str,
    confidence: str,
    source_cases: str = "",
    notes: str = "",
    validated: bool = False,
) -> Path:
    """Append a new rule row to Decision_Rules.

    Returns:
        Path to backup file (for inspection / rollback).

    Raises:
        ValueError: If basic field validation fails.
        KBValidationError: If post-write validation fails.
    """
    if tier.upper() not in S.VALID_TIERS:
        raise ValueError(f"Invalid tier '{tier}'. Valid: {sorted(S.VALID_TIERS)}")
    if not trigger.strip():
        raise ValueError("trigger cannot be empty")
    if not action.strip():
        raise ValueError("action cannot be empty")
    valid_conf = {"HIGH", "MEDIUM", "LOW", "MAXIMA"}
    if confidence.upper() not in valid_conf:
        raise ValueError(f"Invalid confidence '{confidence}'. Valid: {sorted(valid_conf)}")
    canonical = V.extract_rule_id(rule_id)
    if canonical is None:
        raise ValueError(f"'{rule_id}' is not a valid Rule_ID")

    row_data = {
        "Rule_ID":              canonical,
        "Trigger Conditions":   trigger.strip(),
        "Action":               action.strip(),
        "Confidence Required":  confidence.upper(),
        "Source Cases":         source_cases.strip(),
        "Validated?":           "true" if validated else "false",
        "Notes":                notes.strip(),
        "tier":                 tier.upper(),
    }

    def _mutate(wb: Workbook) -> None:
        # Reject duplicate id pre-write (fast fail before backup).
        try:
            find_rule_row(wb, canonical)
            raise ValueError(f"Rule {canonical} already exists")
        except LookupError:
            pass  # not found is the happy path
        append_row(
            wb,
            S.DECISION_RULES_SHEET,
            row_data,
            header_row=S.DECISION_RULES_HEADER_ROW,
        )

    return with_backup_validate(kb_path, f"pre_add_{canonical}", _mutate)


def edit_rule(
    kb_path: Path,
    rule_id: str,
    field: str,
    value: Any,
) -> tuple[Path, Any, Any]:
    """Update a single field on an existing rule. Returns (backup_path, before, after).

    Raises:
        ValueError: If `field` is not editable or invalid.
        LookupError: If rule not found.
        KBValidationError: If post-write validation fails.
    """
    if field == "Rule_ID":
        raise ValueError("Rule_ID is not editable (use delete-rule + add-rule)")
    if field not in S.DECISION_RULES_EXPECTED_HEADERS:
        raise ValueError(
            f"Unknown field '{field}'. Valid: {S.DECISION_RULES_EXPECTED_HEADERS}"
        )

    # Capture before (read-only load is fine for this).
    wb_before = V.load_kb(kb_path)
    row_idx, row_tuple = find_rule_row(wb_before, rule_id)
    col_idx = S.DECISION_RULES_COLS[field] - 1
    before = row_tuple[col_idx] if col_idx < len(row_tuple) else None

    # Tier validation specifically.
    if field == "tier" and str(value).upper() not in S.VALID_TIERS:
        raise ValueError(f"Invalid tier '{value}'. Valid: {sorted(S.VALID_TIERS)}")

    def _mutate(wb: Workbook) -> None:
        idx, _ = find_rule_row(wb, rule_id)
        update_row(
            wb,
            S.DECISION_RULES_SHEET,
            idx,
            {field: value},
            header_row=S.DECISION_RULES_HEADER_ROW,
        )

    canonical = V.extract_rule_id(rule_id) or rule_id
    backup_path = with_backup_validate(
        kb_path, f"pre_edit_{canonical}_{field}", _mutate
    )
    return backup_path, before, value


def delete_rule_soft(
    kb_path: Path,
    rule_id: str,
    reason: str = "",
) -> Path:
    """Mark rule as DEPRECATED in Notes + flip Validated? to false.

    Preserves the row (id remains referenceable). Idempotent: re-deprecating
    appends another tag.
    """
    canonical = V.extract_rule_id(rule_id) or rule_id
    today = datetime.now().strftime("%Y-%m-%d")
    tag = f"[DEPRECATED {today}: {reason}]" if reason else f"[DEPRECATED {today}]"

    # Capture current notes to append.
    wb_before = V.load_kb(kb_path)
    _, row_tuple = find_rule_row(wb_before, canonical)
    notes_col = S.DECISION_RULES_COLS["Notes"] - 1
    current_notes = (
        str(row_tuple[notes_col]) if notes_col < len(row_tuple) and row_tuple[notes_col] is not None else ""
    )
    new_notes = (current_notes + " " + tag).strip() if current_notes else tag

    def _mutate(wb: Workbook) -> None:
        idx, _ = find_rule_row(wb, canonical)
        update_row(
            wb,
            S.DECISION_RULES_SHEET,
            idx,
            {"Notes": new_notes, "Validated?": "false"},
            header_row=S.DECISION_RULES_HEADER_ROW,
        )

    return with_backup_validate(kb_path, f"pre_softdel_{canonical}", _mutate)


def delete_rule_hard(
    kb_path: Path,
    rule_id: str,
    reason: str,
) -> tuple[Path, list[str]]:
    """Physically remove the rule row. Returns (backup_path, ghost_warnings).

    Args:
        reason: Required justification (logged in backup name).

    Returns:
        (backup_path, list of "Case CASE_ID references TR-X" warnings).

    Raises:
        ValueError: If reason is empty.
        LookupError: If rule not found.
    """
    if not reason.strip():
        raise ValueError("--reason required for hard delete")
    canonical = V.extract_rule_id(rule_id) or rule_id

    # Pre-collect ghost refs warnings (cases that cite this rule).
    wb_before = V.load_kb(kb_path)
    ghost_warnings: list[str] = []
    if S.CASES_SHEET in wb_before.sheetnames:
        ws = wb_before[S.CASES_SHEET]
        headers_map: dict[str, int] = {}
        for col_idx, cell in enumerate(ws[S.CASES_HEADER_ROW], start=1):
            if cell.value:
                headers_map[str(cell.value).strip()] = col_idx
        cid_col = headers_map.get("case_id")
        tra_col = headers_map.get("tacit_rules_applied")
        if cid_col and tra_col:
            for row in ws.iter_rows(
                min_row=S.CASES_DATA_START_ROW, values_only=True
            ):
                if not row or len(row) < max(cid_col, tra_col):
                    continue
                cid_val = row[cid_col - 1]
                tra_val = row[tra_col - 1]
                if not cid_val or not tra_val:
                    continue
                cited = {
                    S.RULE_ID_FORMAT.format(num=int(m))
                    for m in S.RULE_ID_REGEX.findall(str(tra_val))
                }
                if canonical in cited:
                    ghost_warnings.append(
                        f"Case {str(cid_val).strip()} cites {canonical}"
                    )

    def _mutate(wb: Workbook) -> None:
        idx, _ = find_rule_row(wb, canonical)
        remove_row(wb, S.DECISION_RULES_SHEET, idx)

    # Hard delete can leave ghost refs intentionally (validator will flag).
    # We disable post-validation strictness for rule_references only.
    backup_path = backup_kb(kb_path, f"pre_harddel_{canonical}")
    try:
        wb = load_workbook(kb_path, data_only=False, read_only=False)
        _mutate(wb)
        save_workbook_atomic(wb, kb_path)
    except Exception:
        shutil.copy2(backup_path, kb_path)
        raise

    # Validate but tolerate rule_references errors that mention the removed id.
    wb_check = V.load_kb(kb_path)
    report = V.validate_all(wb_check)
    fatal: dict[str, list[str]] = {}
    for check, errs in report.items():
        if not errs:
            continue
        if check == "rule_references":
            # Only fatal if errors mention a DIFFERENT rule than canonical
            # (i.e., regression unrelated to our delete).
            unrelated = [e for e in errs if canonical not in e]
            if unrelated:
                fatal[check] = unrelated
        else:
            fatal[check] = errs
    if fatal:
        shutil.copy2(backup_path, kb_path)
        raise KBValidationError(
            f"Hard delete post-validation failed; backup restored. Errors: {fatal}"
        )

    return backup_path, ghost_warnings


def merge_rules(
    kb_path: Path,
    rule_ids: list[str],
    into: Optional[str],
    reason: str,
) -> Path:
    """Combine triggers + actions of `rule_ids` into `into` (lowest by default).

    The "source" rules (not `into`) get a DEPRECATED tag appended to Notes.
    The target rule's Trigger, Action, and Notes are concatenated.

    Args:
        rule_ids: List of canonical or raw ids to merge (>=2).
        into: Target rule id (must be one of `rule_ids`). If None, picks lowest.
        reason: Justification (appended to Notes).

    Raises:
        ValueError: If fewer than 2 ids or `into` not in `rule_ids`.
    """
    if len(rule_ids) < 2:
        raise ValueError("merge-rules requires at least 2 rule ids")
    canonicals = [V.extract_rule_id(r) for r in rule_ids]
    if any(c is None for c in canonicals):
        raise ValueError(f"Invalid rule ids: {rule_ids}")
    canonicals = [c for c in canonicals if c]  # narrow type

    if into is None:
        target = min(canonicals, key=lambda x: int(S.RULE_ID_REGEX.search(x).group(1)))
    else:
        target_canonical = V.extract_rule_id(into)
        if target_canonical not in canonicals:
            raise ValueError(f"--into '{into}' must be one of {rule_ids}")
        target = target_canonical

    sources = [c for c in canonicals if c != target]
    today = datetime.now().strftime("%Y-%m-%d")
    merge_tag = f"[MERGED {today}: {'+'.join(canonicals)} -> {target}. {reason}]"

    # Collect data from all rules (read-only).
    wb_before = V.load_kb(kb_path)
    rule_data: dict[str, dict[str, Any]] = {}
    for rid in canonicals:
        _, row = find_rule_row(wb_before, rid)
        rule_data[rid] = {
            header: row[idx - 1] if idx - 1 < len(row) else None
            for header, idx in S.DECISION_RULES_COLS.items()
        }

    # Build target's new fields.
    target_data = rule_data[target]
    new_trigger = "; ".join(
        str(rule_data[r].get("Trigger Conditions") or "").strip()
        for r in canonicals
        if rule_data[r].get("Trigger Conditions")
    )
    new_action = "; ".join(
        str(rule_data[r].get("Action") or "").strip()
        for r in canonicals
        if rule_data[r].get("Action")
    )
    target_notes_current = str(target_data.get("Notes") or "").strip()
    target_notes_new = (
        target_notes_current + " " + merge_tag
    ).strip() if target_notes_current else merge_tag

    def _mutate(wb: Workbook) -> None:
        target_idx, _ = find_rule_row(wb, target)
        update_row(
            wb,
            S.DECISION_RULES_SHEET,
            target_idx,
            {
                "Trigger Conditions": new_trigger,
                "Action": new_action,
                "Notes": target_notes_new,
            },
            header_row=S.DECISION_RULES_HEADER_ROW,
        )
        # Mark sources DEPRECATED.
        for src in sources:
            src_idx, src_tuple = find_rule_row(wb, src)
            src_notes_col = S.DECISION_RULES_COLS["Notes"] - 1
            src_notes_current = (
                str(src_tuple[src_notes_col])
                if src_notes_col < len(src_tuple) and src_tuple[src_notes_col]
                else ""
            )
            src_tag = f"[DEPRECATED {today}: merged into {target}. {reason}]"
            src_notes_new = (src_notes_current + " " + src_tag).strip() if src_notes_current else src_tag
            update_row(
                wb,
                S.DECISION_RULES_SHEET,
                src_idx,
                {"Notes": src_notes_new, "Validated?": "false"},
                header_row=S.DECISION_RULES_HEADER_ROW,
            )

    return with_backup_validate(kb_path, f"pre_merge_into_{target}", _mutate)


# ── Version bump ──────────────────────────────────────────────────────────
def bump_version(
    kb_path: Path,
    from_version: str,
    to_version: str,
) -> tuple[Path, Path]:
    """Backup + rename KB file + update kb_loader.py and kb_schema.py references.

    Returns:
        (backup_path, new_kb_path).

    Raises:
        ValueError: If versions invalid or current KB doesn't match `from_version`.
        FileNotFoundError: If kb_path missing.
        KBValidationError: If pre-bump validation fails.
    """
    version_re = re.compile(r"^v\d+\.\d+$")
    if not version_re.match(from_version) or not version_re.match(to_version):
        raise ValueError(
            f"Invalid versions. Expected vN.M (got from={from_version}, to={to_version})"
        )

    if not kb_path.exists():
        raise FileNotFoundError(f"KB not found: {kb_path}")

    # Source filename must contain from_version.
    if from_version not in kb_path.name:
        raise ValueError(
            f"KB filename '{kb_path.name}' does not contain '{from_version}'"
        )

    # Pre-validate current KB (no bump if broken).
    wb = V.load_kb(kb_path)
    report = V.validate_all(wb)
    if any(report.values()):
        bad = {k: v for k, v in report.items() if v}
        raise KBValidationError(
            f"Refusing to bump: current KB has validation errors: {bad}"
        )

    # Backup with FINAL tag.
    backup_path = backup_kb(kb_path, f"{from_version}_FINAL")

    # New filename.
    new_name = kb_path.name.replace(from_version, to_version)
    new_path = kb_path.parent / new_name
    if new_path.exists():
        raise FileExistsError(f"Target KB already exists: {new_path}")

    # Rename atomically.
    os.replace(kb_path, new_path)

    # Update references in code.
    _replace_in_file(
        S.REPO_ROOT / "llm_engine_eolo" / "llm_engine" / "kb_loader.py",
        from_version,
        to_version,
    )
    _replace_in_file(
        S.REPO_ROOT / "tools" / "kb_schema.py",
        from_version,
        to_version,
    )

    return backup_path, new_path


def _replace_in_file(file_path: Path, old: str, new: str) -> None:
    """In-place string replacement. No-op if file missing or `old` not found."""
    if not file_path.exists():
        return
    content = file_path.read_text(encoding="utf-8")
    if old not in content:
        return
    file_path.write_text(content.replace(old, new), encoding="utf-8")
