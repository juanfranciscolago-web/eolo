"""KB Editor CLI.

Phase 1 (read-only):
    python tools/kb_editor.py list-rules
    python tools/kb_editor.py next-id
    python tools/kb_editor.py show-rule TR-Juan-043
    python tools/kb_editor.py validate
    python tools/kb_editor.py list-cases
    python tools/kb_editor.py stats

Phase 2 (UP-1.2 fase 2 — write):
    python tools/kb_editor.py add-rule --tier ... --trigger ... --action ...
    python tools/kb_editor.py edit-rule TR-Juan-043 --field Notes --value "..."
    python tools/kb_editor.py delete-rule TR-Juan-018 [--hard --reason "..."]
    python tools/kb_editor.py merge-rules TR-Juan-022 TR-Juan-031 --into TR-Juan-022 --reason "..."
    python tools/kb_editor.py bump-version --from v1.2 --to v1.3

All commands accept `--kb-path PATH` to target a non-default workbook.
Write commands create a timestamped backup in `backups/` and auto-restore
if post-write validation fails.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path
from typing import Iterator

from openpyxl.workbook.workbook import Workbook

# Ensure `tools` is importable when this file is invoked as a script from any cwd.
_HERE = Path(__file__).resolve().parent
if str(_HERE.parent) not in sys.path:
    sys.path.insert(0, str(_HERE.parent))

from tools import kb_schema as S  # noqa: E402
from tools import kb_validators as V  # noqa: E402
from tools import kb_writer as W  # noqa: E402


# ── Iterators ─────────────────────────────────────────────────────────────
def _iter_rule_rows(wb: Workbook) -> Iterator[tuple[int, str, tuple]]:
    """Yield (row_idx, canonical_rule_id, row_tuple) for every rule.

    Args:
        wb: Loaded workbook.

    Yields:
        (1-indexed row number, normalised TR-Juan-NNN, raw row tuple).

    Raises:
        KeyError: If the Decision_Rules sheet is missing.
    """
    if S.DECISION_RULES_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{S.DECISION_RULES_SHEET}' missing from workbook")
    ws = wb[S.DECISION_RULES_SHEET]
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=S.DECISION_RULES_DATA_START_ROW, values_only=True),
        start=S.DECISION_RULES_DATA_START_ROW,
    ):
        if not row or row[0] is None:
            continue
        if not str(row[0]).strip().startswith("TR-"):
            continue
        rid = V.extract_rule_id(row[0])
        if rid is None:
            continue
        yield row_idx, rid, row


def _iter_case_rows(wb: Workbook) -> Iterator[tuple[int, str, dict[str, object]]]:
    """Yield (row_idx, case_id, {header: value}) for every case.

    Args:
        wb: Loaded workbook.

    Yields:
        (1-indexed row number, case_id, dict mapping header to cell value).

    Raises:
        KeyError: If the Cases sheet is missing.
    """
    if S.CASES_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{S.CASES_SHEET}' missing from workbook")
    ws = wb[S.CASES_SHEET]
    headers: list[str] = [
        (str(c.value).strip() if c.value is not None else "")
        for c in ws[S.CASES_HEADER_ROW]
    ]
    cid_idx = headers.index("case_id") if "case_id" in headers else 0
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=S.CASES_DATA_START_ROW, values_only=True),
        start=S.CASES_DATA_START_ROW,
    ):
        if not row or cid_idx >= len(row) or row[cid_idx] is None:
            continue
        case_id = str(row[cid_idx]).strip()
        if not case_id:
            continue
        record = {
            h: row[i] if i < len(row) else None
            for i, h in enumerate(headers)
            if h
        }
        yield row_idx, case_id, record


# ── Commands ──────────────────────────────────────────────────────────────
def cmd_list_rules(wb: Workbook) -> int:
    """Print every Rule_ID with its tier, one per line.

    Args:
        wb: Loaded workbook.

    Returns:
        Process exit code (0 on success).
    """
    tier_col_idx = S.DECISION_RULES_COLS["tier"] - 1
    count = 0
    for _, rid, row in _iter_rule_rows(wb):
        tier = ""
        if len(row) > tier_col_idx and row[tier_col_idx] is not None:
            tier = str(row[tier_col_idx]).strip().upper()
        print(f"{rid}  [{tier or '?'}]")
        count += 1
    print(f"\nTotal: {count} rules")
    return 0


def cmd_next_id(wb: Workbook) -> int:
    """Print the next free TR-Juan-NNN id.

    The "next free" id is one greater than the largest numeric component
    currently present in the Decision_Rules sheet.

    Args:
        wb: Loaded workbook.

    Returns:
        0 on success.
    """
    max_num = 0
    for _, rid, _ in _iter_rule_rows(wb):
        m = S.RULE_ID_REGEX.search(rid)
        if m:
            max_num = max(max_num, int(m.group(1)))
    print(S.RULE_ID_FORMAT.format(num=max_num + 1))
    return 0


def cmd_show_rule(wb: Workbook, rule_id_arg: str) -> int:
    """Print every column of the rule whose canonical id matches `rule_id_arg`.

    Args:
        wb: Loaded workbook.
        rule_id_arg: The id to look up (any form containing TR-Juan-NNN).

    Returns:
        0 if found, 1 if no rule matches.
    """
    target = V.extract_rule_id(rule_id_arg)
    if target is None:
        print(f"ERROR: '{rule_id_arg}' is not a valid Rule_ID", file=sys.stderr)
        return 1

    for row_idx, rid, row in _iter_rule_rows(wb):
        if rid != target:
            continue
        print(f"Rule {rid} (row {row_idx})")
        print("─" * 60)
        for header, col_pos in S.DECISION_RULES_COLS.items():
            idx = col_pos - 1
            value = row[idx] if idx < len(row) else None
            display = "" if value is None else str(value)
            print(f"  {header:>22}: {display}")
        return 0

    print(f"ERROR: rule {target} not found", file=sys.stderr)
    return 1


def cmd_validate(wb: Workbook) -> int:
    """Run every validator and report the result.

    Args:
        wb: Loaded workbook.

    Returns:
        0 if every check passed, 1 otherwise.
    """
    report = V.validate_all(wb)
    total_errors = sum(len(errs) for errs in report.values())

    for check, errors in report.items():
        if errors:
            print(f"[FAIL] {check}: {len(errors)} issue(s)")
            for e in errors:
                print(f"   - {e}")
        else:
            print(f"[ OK ] {check}")

    print()
    if total_errors == 0:
        print("✓ All validations passed")
        return 0
    print(f"✗ {total_errors} validation issue(s) across {sum(1 for e in report.values() if e)} check(s)")
    return 1


def cmd_list_cases(wb: Workbook) -> int:
    """Print every case_id and its case_quality.

    Args:
        wb: Loaded workbook.

    Returns:
        0 on success.
    """
    count = 0
    for _, case_id, record in _iter_case_rows(wb):
        quality_raw = record.get("case_quality")
        quality = str(quality_raw).strip().upper() if quality_raw else "?"
        print(f"{case_id}  [{quality}]")
        count += 1
    print(f"\nTotal: {count} cases")
    return 0


def cmd_stats(wb: Workbook) -> int:
    """Print rules-per-tier and cases-per-quality counts.

    Args:
        wb: Loaded workbook.

    Returns:
        0 on success.
    """
    tier_col_idx = S.DECISION_RULES_COLS["tier"] - 1
    tier_counts: Counter[str] = Counter()
    rule_total = 0
    for _, _, row in _iter_rule_rows(wb):
        rule_total += 1
        tier = ""
        if len(row) > tier_col_idx and row[tier_col_idx] is not None:
            tier = str(row[tier_col_idx]).strip().upper()
        tier_counts[tier or "?"] += 1

    quality_counts: Counter[str] = Counter()
    case_total = 0
    for _, _, record in _iter_case_rows(wb):
        case_total += 1
        q_raw = record.get("case_quality")
        q = str(q_raw).strip().upper() if q_raw else "?"
        quality_counts[q] += 1

    print(f"Rules: {rule_total}")
    for tier in S.TIER_PRIORITY_ORDER:
        print(f"  {tier:>14}: {tier_counts.get(tier, 0)}")
    extras = sorted(set(tier_counts) - set(S.TIER_PRIORITY_ORDER))
    for tier in extras:
        print(f"  {tier:>14}: {tier_counts[tier]}  (unexpected)")

    print(f"\nCases: {case_total}")
    for q in ("GOLD", "SILVER", "BRONZE"):
        print(f"  {q:>14}: {quality_counts.get(q, 0)}")
    extras = sorted(set(quality_counts) - {"GOLD", "SILVER", "BRONZE"})
    for q in extras:
        print(f"  {q:>14}: {quality_counts[q]}  (unexpected)")
    return 0


# ── Write commands (UP-1.2 fase 2) ────────────────────────────────────────
def _confirm(prompt: str, assume_yes: bool = False) -> bool:
    """Ask yes/no on stdin. Returns True if confirmed."""
    if assume_yes:
        return True
    try:
        resp = input(f"{prompt} [y/N]: ").strip().lower()
    except EOFError:
        return False
    return resp in {"y", "yes"}


def _prompt(prompt: str, default: str | None = None, multiline: bool = False) -> str:
    """Read a value from stdin, with optional default. multiline=True means
    accept until empty line."""
    if multiline:
        print(f"{prompt} (end with empty line):")
        lines: list[str] = []
        try:
            while True:
                line = input()
                if not line:
                    break
                lines.append(line)
        except EOFError:
            pass
        result = "\n".join(lines).strip()
        return result if result else (default or "")
    suffix = f" [{default}]" if default else ""
    try:
        val = input(f"{prompt}{suffix}: ").strip()
    except EOFError:
        val = ""
    return val if val else (default or "")


def cmd_add_rule(kb_path: Path, args: argparse.Namespace) -> int:
    """Append a new rule to Decision_Rules.

    Two modes:
      - CLI: --tier, --trigger, --action, --confidence all provided.
      - Interactive: prompts for each field if any required arg missing.
    """
    wb_read = V.load_kb(kb_path)
    proposed_id = args.rule_id or W.compute_next_rule_id(wb_read)

    # Interactive mode if required CLI args missing.
    if not (args.tier and args.trigger and args.action and args.confidence):
        print(f"Proposed Rule_ID: {proposed_id}")
        rule_id = _prompt("Rule_ID", default=proposed_id)
        tier = args.tier or _prompt(
            f"Tier (one of: {sorted(S.VALID_TIERS)})", default="TACTICAL_PLUS"
        )
        trigger = args.trigger or _prompt("Trigger Conditions", multiline=True)
        action = args.action or _prompt("Action", multiline=True)
        confidence = args.confidence or _prompt(
            "Confidence (HIGH/MEDIUM/LOW/MAXIMA)", default="MEDIUM"
        )
        source_cases = args.source_cases or _prompt("Source Cases (optional)")
        notes = args.notes or _prompt("Notes (optional)")
        validated_str = _prompt("Validated? (true/false)", default="false")
        validated = validated_str.lower() == "true"
    else:
        rule_id = args.rule_id or proposed_id
        tier = args.tier
        trigger = args.trigger
        action = args.action
        confidence = args.confidence
        source_cases = args.source_cases or ""
        notes = args.notes or ""
        validated = args.validated

    # Preview + confirm.
    print()
    print("─" * 60)
    print(f"New rule preview:")
    print(f"  Rule_ID:    {rule_id}")
    print(f"  tier:       {tier.upper()}")
    print(f"  Confidence: {confidence.upper()}")
    print(f"  Trigger:    {trigger[:80]}{'...' if len(trigger) > 80 else ''}")
    print(f"  Action:     {action[:80]}{'...' if len(action) > 80 else ''}")
    print(f"  Source:     {source_cases}")
    print(f"  Notes:      {notes[:80]}{'...' if len(notes) > 80 else ''}")
    print(f"  Validated?: {validated}")
    print("─" * 60)
    if not _confirm("Proceed with add?", assume_yes=args.yes):
        print("Aborted.")
        return 1

    try:
        backup = W.add_rule(
            kb_path=kb_path,
            rule_id=rule_id,
            tier=tier,
            trigger=trigger,
            action=action,
            confidence=confidence,
            source_cases=source_cases,
            notes=notes,
            validated=validated,
        )
    except (ValueError, W.KBValidationError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    print(f"✅ Added {V.extract_rule_id(rule_id)} to KB. Backup: {backup}")
    return 0


def cmd_edit_rule(kb_path: Path, args: argparse.Namespace) -> int:
    """Update a single field on an existing rule."""
    try:
        wb_read = V.load_kb(kb_path)
        row_idx, row_tuple = W.find_rule_row(wb_read, args.rule_id)
    except LookupError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    col_idx = S.DECISION_RULES_COLS.get(args.field)
    if col_idx is None:
        print(
            f"ERROR: unknown field '{args.field}'. "
            f"Valid: {S.DECISION_RULES_EXPECTED_HEADERS}",
            file=sys.stderr,
        )
        return 1
    current = row_tuple[col_idx - 1] if col_idx - 1 < len(row_tuple) else None
    print(f"Editing {V.extract_rule_id(args.rule_id)} (row {row_idx}), field '{args.field}'")
    print(f"  Before: {current}")
    print(f"  After:  {args.value}")
    if not _confirm("Proceed?", assume_yes=args.yes):
        print("Aborted.")
        return 1

    try:
        backup, before, after = W.edit_rule(
            kb_path=kb_path,
            rule_id=args.rule_id,
            field=args.field,
            value=args.value,
        )
    except (ValueError, LookupError, W.KBValidationError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    print(f"✅ Updated. Backup: {backup}")
    return 0


def cmd_delete_rule(kb_path: Path, args: argparse.Namespace) -> int:
    """Soft (DEPRECATED) or --hard delete a rule."""
    mode = "HARD" if args.hard else "SOFT"
    print(f"Delete {V.extract_rule_id(args.rule_id) or args.rule_id} ({mode})")
    if args.hard:
        if not args.reason:
            print("ERROR: --reason required for --hard delete", file=sys.stderr)
            return 1
        print(f"  Reason: {args.reason}")
    if not _confirm("Proceed?", assume_yes=args.yes):
        print("Aborted.")
        return 1

    try:
        if args.hard:
            backup, ghosts = W.delete_rule_hard(
                kb_path, args.rule_id, args.reason or ""
            )
            print(f"✅ Hard-deleted. Backup: {backup}")
            if ghosts:
                print(f"⚠️ Ghost references in Cases:")
                for g in ghosts:
                    print(f"   - {g}")
        else:
            backup = W.delete_rule_soft(
                kb_path, args.rule_id, args.reason or ""
            )
            print(f"✅ Soft-deleted (DEPRECATED tag). Backup: {backup}")
    except (ValueError, LookupError, W.KBValidationError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    return 0


def cmd_merge_rules(kb_path: Path, args: argparse.Namespace) -> int:
    """Combine 2+ rules into one. Sources marked DEPRECATED."""
    if len(args.rule_ids) < 2:
        print("ERROR: merge-rules requires at least 2 rule ids", file=sys.stderr)
        return 1
    if not args.reason:
        print("ERROR: --reason required", file=sys.stderr)
        return 1

    target_preview = args.into or "lowest"
    print(f"Merging {args.rule_ids} into {target_preview}")
    print(f"  Reason: {args.reason}")
    if not _confirm("Proceed?", assume_yes=args.yes):
        print("Aborted.")
        return 1

    try:
        backup = W.merge_rules(
            kb_path=kb_path,
            rule_ids=list(args.rule_ids),
            into=args.into,
            reason=args.reason,
        )
    except (ValueError, LookupError, W.KBValidationError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    print(f"✅ Merged. Backup: {backup}")
    return 0


def cmd_bump_version(kb_path: Path, args: argparse.Namespace) -> int:
    """Backup + rename KB + (optionally) update kb_loader.py + kb_schema.py references."""
    is_default = kb_path.resolve() == S.DEFAULT_KB_PATH.resolve()
    print(f"Bumping KB: {args.from_version} → {args.to_version}")
    print(f"  Current path: {kb_path}")
    if is_default:
        print(f"  Update code refs: True (canonical KB)")
    else:
        print(f"  Update code refs: False (override path → skip patching kb_schema.py / kb_loader.py)")
    if not _confirm("Proceed?", assume_yes=args.yes):
        print("Aborted.")
        return 1

    try:
        backup, new_path = W.bump_version(
            kb_path=kb_path,
            from_version=args.from_version,
            to_version=args.to_version,
            update_code_refs=is_default,
        )
    except (ValueError, FileNotFoundError, FileExistsError, W.KBValidationError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    print(f"✅ Bumped KB to {args.to_version}. Backup: {backup}")
    print(f"   New KB:  {new_path}")
    print()
    if is_default:
        print("⚠️ REMINDER: redeploy LLM Engine para que cargue la nueva KB.")
        print("   gcloud builds submit --config llm_engine_eolo/cloudbuild.yaml .")
    else:
        print("ℹ️  Override path: no redeploy needed. tools/kb_schema.py NOT touched.")
    return 0


# ── Entrypoint ────────────────────────────────────────────────────────────
def build_parser() -> argparse.ArgumentParser:
    """Build the argparse CLI.

    Returns:
        Configured ArgumentParser with subcommands wired to handlers.
    """
    parser = argparse.ArgumentParser(
        prog="kb_editor",
        description="EOLO ThetaHarvest KB editor (phase 1 read + phase 2 write)",
    )
    parser.add_argument(
        "--kb-path",
        type=Path,
        default=S.DEFAULT_KB_PATH,
        help=f"Path to the KB .xlsx (default: {S.DEFAULT_KB_PATH})",
    )

    sub = parser.add_subparsers(dest="command", required=True)
    # Phase 1: read-only.
    sub.add_parser("list-rules", help="List Rule_IDs with their tier")
    sub.add_parser("next-id", help="Print the next free TR-Juan-NNN")
    p_show = sub.add_parser("show-rule", help="Show a rule's full row")
    p_show.add_argument("rule_id", help="Rule id (e.g. TR-Juan-043)")
    sub.add_parser("validate", help="Run schema and integrity validators")
    sub.add_parser("list-cases", help="List case_ids with their case_quality")
    sub.add_parser("stats", help="Count rules per tier and cases per quality")

    # Phase 2: write commands.
    p_add = sub.add_parser("add-rule", help="Append a new rule (interactive if args missing)")
    p_add.add_argument("--rule-id", default=None, help="Override auto-computed id")
    p_add.add_argument("--tier", default=None, help="One of " + ", ".join(sorted(S.VALID_TIERS)))
    p_add.add_argument("--trigger", default=None, help="Trigger Conditions")
    p_add.add_argument("--action", default=None, help="Action")
    p_add.add_argument("--confidence", default=None, help="HIGH/MEDIUM/LOW/MAXIMA")
    p_add.add_argument("--source-cases", default=None, help="Cited case_ids")
    p_add.add_argument("--notes", default=None, help="Free-form notes")
    p_add.add_argument("--validated", action="store_true", help="Mark as validated")
    p_add.add_argument("-y", "--yes", action="store_true", help="Skip confirm prompt")

    p_edit = sub.add_parser("edit-rule", help="Update a single field on a rule")
    p_edit.add_argument("rule_id", help="Rule id (e.g. TR-Juan-043)")
    p_edit.add_argument(
        "--field",
        required=True,
        help=f"One of: {[h for h in S.DECISION_RULES_EXPECTED_HEADERS if h != 'Rule_ID']}",
    )
    p_edit.add_argument("--value", required=True, help="New value")
    p_edit.add_argument("-y", "--yes", action="store_true", help="Skip confirm prompt")

    p_del = sub.add_parser("delete-rule", help="Soft delete (DEPRECATED) or --hard physical")
    p_del.add_argument("rule_id", help="Rule id")
    p_del.add_argument("--hard", action="store_true", help="Physical remove (requires --reason)")
    p_del.add_argument("--reason", default="", help="Justification (required if --hard)")
    p_del.add_argument("-y", "--yes", action="store_true", help="Skip confirm prompt")

    p_merge = sub.add_parser("merge-rules", help="Combine 2+ rules; sources DEPRECATED")
    p_merge.add_argument("rule_ids", nargs="+", help="Rule ids to merge (>=2)")
    p_merge.add_argument("--into", default=None, help="Target id (default: lowest)")
    p_merge.add_argument("--reason", required=True, help="Justification")
    p_merge.add_argument("-y", "--yes", action="store_true", help="Skip confirm prompt")

    p_bump = sub.add_parser("bump-version", help="Backup + rename KB + update refs")
    p_bump.add_argument("--from", dest="from_version", required=True, help="e.g. v1.2")
    p_bump.add_argument("--to", dest="to_version", required=True, help="e.g. v1.3")
    p_bump.add_argument("-y", "--yes", action="store_true", help="Skip confirm prompt")

    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entrypoint.

    Args:
        argv: Optional argument list; defaults to sys.argv[1:].

    Returns:
        Process exit code.
    """
    args = build_parser().parse_args(argv)

    # Write commands handle their own KB loading (open/save cycles).
    write_handlers = {
        "add-rule":     lambda: cmd_add_rule(args.kb_path, args),
        "edit-rule":    lambda: cmd_edit_rule(args.kb_path, args),
        "delete-rule":  lambda: cmd_delete_rule(args.kb_path, args),
        "merge-rules":  lambda: cmd_merge_rules(args.kb_path, args),
        "bump-version": lambda: cmd_bump_version(args.kb_path, args),
    }
    if args.command in write_handlers:
        try:
            return write_handlers[args.command]()
        except FileNotFoundError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            return 2

    # Read commands load workbook once.
    try:
        wb = V.load_kb(args.kb_path)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2

    handlers = {
        "list-rules": lambda: cmd_list_rules(wb),
        "next-id":    lambda: cmd_next_id(wb),
        "show-rule":  lambda: cmd_show_rule(wb, args.rule_id),
        "validate":   lambda: cmd_validate(wb),
        "list-cases": lambda: cmd_list_cases(wb),
        "stats":      lambda: cmd_stats(wb),
    }
    try:
        return handlers[args.command]()
    except KeyError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())
