"""
KB Loader v2 - Carga el Knowledge Base de Juan desde Excel.

CAMBIOS v2 (fix bugs Juan 27-may):
- Lee 'tier' de columna explícita (col 8), no de string matching
- Soporta nuevo tier "TACTICAL_PLUS" (reglas con ⭐ pero no MAESTRA)
- Valida que reglas referenciadas en casos existan en Decision_Rules
- Detecta y reporta reglas fantasma

Procesa:
- Decision_Rules sheet (47 tacit rules)
- Cases sheet (6 casos documentados)
- Juan_Trading_Thesis (tesis dual-engine)

Tiers:
- AXIOMA (foundational)
- PROHIBITIVA (hard rules)
- MAESTRA (core rules)
- PROTOCOLO (workflow)
- TACTICAL_PLUS (situational pero alta prioridad)
- TACTICAL (situational baseline)
"""
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from typing import List, Optional, Set
from pathlib import Path
import logging
import re

logger = logging.getLogger(__name__)


VALID_TIERS = {"AXIOMA", "PROHIBITIVA", "MAESTRA", "PROTOCOLO", "TACTICAL_PLUS", "TACTICAL"}


class TacitRule(BaseModel):
    """Una regla tácita del sistema de Juan."""
    rule_id: str
    trigger: str
    action: str
    priority: str = "MEDIUM"
    tier: str = "TACTICAL"
    source: str = ""
    validation_status: str = "PENDING"
    notes: str = ""

    def normalized_id(self) -> str:
        """Devuelve solo el rule_id base sin markers (ej TR-Juan-010)."""
        m = re.match(r'(TR-Juan-\d+)', self.rule_id)
        return m.group(1) if m else self.rule_id


class Case(BaseModel):
    """Un caso documentado del KB."""
    case_id: str
    date: str
    time_analysis: str = ""
    ticker: str = "SPY"
    setup_type: str = ""
    juan_action: str = ""
    juan_confidence: int = 0
    juan_reasoning: str = ""
    tacit_rules_applied: str = ""
    outcome: Optional[str] = None
    pnl_pct: Optional[str] = None
    lesson_learned: str = ""
    case_quality: str = "SILVER"
    rag_tags: str = ""

    def get_referenced_rules(self) -> Set[str]:
        """Extrae los rule_ids referenciados en tacit_rules_applied."""
        return set(re.findall(r'TR-Juan-\d+', self.tacit_rules_applied or ""))


class KBLoader:
    """Carga y consulta del Knowledge Base v2."""

    def __init__(self, excel_path: str, validate_references: bool = True):
        path = Path(excel_path)
        if not path.exists():
            raise FileNotFoundError(f"KB Excel not found at {excel_path}")

        logger.info(f"Loading KB from {excel_path}")
        self.wb = load_workbook(excel_path, data_only=True)
        self.rules: List[TacitRule] = self._load_rules()
        self.cases: List[Case] = self._load_cases()

        self._rule_index = {r.normalized_id(): r for r in self.rules}

        if validate_references:
            self._validate_case_references()

        logger.info(
            f"KB loaded: {len(self.rules)} rules, "
            f"{len(self.cases)} cases"
        )
        logger.info(f"Tier distribution: {self._tier_distribution()}")

    def _load_rules(self) -> List[TacitRule]:
        if 'Decision_Rules' not in self.wb.sheetnames:
            logger.warning("Decision_Rules sheet not found")
            return []

        ws = self.wb['Decision_Rules']
        rules = []

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
        deprecated_by_idx = header_row.index("Deprecated_By") if "Deprecated_By" in header_row else None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row or not row[0]:
                continue

            rule_id = str(row[0]).strip()
            if not rule_id.startswith("TR-"):
                continue

            if deprecated_by_idx is not None and deprecated_by_idx < len(row):
                dep_by = row[deprecated_by_idx]
                if dep_by and str(dep_by).strip():
                    logger.info(f"Skipping {rule_id} (deprecated_by={dep_by})")
                    continue

            tier_from_col = None
            if len(row) >= 8 and row[7]:
                tier_str = str(row[7]).strip().upper()
                if tier_str in VALID_TIERS:
                    tier_from_col = tier_str

            if not tier_from_col:
                tier_from_col = self._infer_tier_from_id(rule_id)
                logger.warning(
                    f"Rule {rule_id}: no tier column, inferring '{tier_from_col}' from ID. "
                    f"Update Excel with explicit tier column."
                )

            rule = TacitRule(
                rule_id=rule_id,
                trigger=str(row[1] or "").strip(),
                action=str(row[2] or "").strip(),
                priority=str(row[3] or "MEDIUM").strip().upper(),
                tier=tier_from_col,
                source=str(row[4] or "").strip() if len(row) > 4 else "",
                validation_status=str(row[5] or "PENDING").strip() if len(row) > 5 else "PENDING",
                notes=str(row[6] or "").strip() if len(row) > 6 else "",
            )
            rules.append(rule)

        return rules

    def _infer_tier_from_id(self, rule_id: str) -> str:
        if "AXIOMA" in rule_id:
            return "AXIOMA"
        if "PROHIBITIVA" in rule_id:
            return "PROHIBITIVA"
        if "PROTOCOLO" in rule_id:
            return "PROTOCOLO"
        if "MAESTRA" in rule_id:
            return "MAESTRA"
        if "⭐" in rule_id:
            return "TACTICAL_PLUS"
        return "TACTICAL"

    def _load_cases(self) -> List[Case]:
        if 'Cases' not in self.wb.sheetnames:
            logger.warning("Cases sheet not found")
            return []

        ws = self.wb['Cases']
        cases = []

        headers = {}
        for col_idx, cell in enumerate(ws[2], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue

            case_id = str(row[0]).strip()
            if not case_id.startswith("2026") and not case_id.startswith("2025"):
                continue

            def get(field_name, default=""):
                if field_name in headers:
                    idx = headers[field_name] - 1
                    if idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                return default

            try:
                confidence_str = get("j_confidence", "0")
                confidence = int(float(confidence_str)) if confidence_str else 0
            except (ValueError, TypeError):
                confidence = 0

            case = Case(
                case_id=case_id,
                date=get("date"),
                time_analysis=get("time_analysis"),
                ticker=get("ticker", "SPY"),
                setup_type=get("session_label") or get("price_action_pattern"),
                juan_action=get("j_action"),
                juan_confidence=confidence,
                juan_reasoning=get("juan_saw_extra"),
                tacit_rules_applied=get("tacit_rules_applied"),
                outcome=get("pnl_pct"),
                lesson_learned=get("lesson_learned"),
                case_quality=get("case_quality", "SILVER"),
                rag_tags=get("rag_tags"),
            )
            cases.append(case)

        return cases

    def _validate_case_references(self):
        """Valida que reglas referenciadas en casos existan."""
        all_rule_ids = set(self._rule_index.keys())
        ghost_refs = {}

        for case in self.cases:
            referenced = case.get_referenced_rules()
            missing = referenced - all_rule_ids
            if missing:
                ghost_refs[case.case_id] = missing

        if ghost_refs:
            logger.error("=" * 60)
            logger.error("GHOST RULES DETECTED — referenced in cases but not defined:")
            for case_id, ghosts in ghost_refs.items():
                logger.error(f"  Case {case_id}: {sorted(ghosts)}")
            logger.error("=" * 60)
        else:
            logger.info("✓ All rule references in cases are valid")

    def _tier_distribution(self) -> dict:
        dist = {}
        for r in self.rules:
            dist[r.tier] = dist.get(r.tier, 0) + 1
        return dist

    def get_rules_by_tier(self, tier: str) -> List[TacitRule]:
        return [r for r in self.rules if r.tier == tier]

    def get_rule_by_id(self, rule_id: str) -> Optional[TacitRule]:
        m = re.match(r'(TR-Juan-\d+)', rule_id)
        if not m:
            return None
        return self._rule_index.get(m.group(1))

    # Sprint ANTI-HALLUCINATION-FIX: validators externos usan estos sets.
    def get_all_rule_ids(self) -> set:
        """Returns set of canonical rule_ids (TR-Juan-XXX) loaded en KB.

        Importante: algunos rule_ids del Excel tienen sufijos decorativos
        como '⭐ MAESTRA' o '⭐⭐ PROHIBITIVA'. Usamos self._rule_index.keys()
        que ya están canonicalizados (sin sufijo) — el mismo set que usa
        get_rule_by_id() para lookups.
        """
        return set(self._rule_index.keys())

    def get_all_case_ids(self) -> set:
        """Returns set of all case_ids in the loaded KB."""
        out = set()
        for c in self.cases:
            cid = getattr(c, "case_id", None)
            if cid:
                # Tomar el primer token sin espacios (defensivo si vienen sufijos)
                token = str(cid).strip().split()[0]
                if token:
                    out.add(token)
        return out

    def get_priority_rules(self) -> List[TacitRule]:
        """Devuelve reglas en orden de prioridad para el prompt."""
        order = ["AXIOMA", "PROHIBITIVA", "MAESTRA", "PROTOCOLO", "TACTICAL_PLUS", "TACTICAL"]
        result = []
        for tier in order:
            result.extend(self.get_rules_by_tier(tier))
        return result

    def get_similar_cases(self, setup_keywords: List[str], top_k: int = 3) -> List[Case]:
        scored = []
        for case in self.cases:
            score = 0
            text = f"{case.setup_type} {case.juan_action} {case.rag_tags}".lower()
            for kw in setup_keywords:
                if kw.lower() in text:
                    score += 1
            if score > 0:
                scored.append((score, case))

        scored.sort(key=lambda x: x[0], reverse=True)
        return [case for _, case in scored[:top_k]]

    def balanced_get_similar_cases(self, snapshot, top_k: int = 3) -> List[Case]:
        """Sprint 8 (T7): balanced sampling por régimen para evitar RAG bias.

        Devuelve al menos 1 case del MISMO régimen + 1 de DIFERENTE + 1 best overall.
        Legacy get_similar_cases sigue disponible sin cambios.
        """
        keywords = snapshot.get_setup_keywords() if hasattr(snapshot, "get_setup_keywords") else []
        current_regime = getattr(snapshot, "gamma_regime_v2", "unknown")

        all_matches = []
        for case in self.cases:
            text = f"{case.setup_type} {case.juan_action} {case.rag_tags}".lower()
            score = sum(1 for kw in keywords if kw.lower() in text)
            if score > 0:
                all_matches.append((score, case))

        by_regime: dict = {}
        for score, case in all_matches:
            regime = getattr(case, "regime", "unknown")
            by_regime.setdefault(regime, []).append((score, case))

        selected: List[Case] = []
        same = by_regime.get(current_regime, [])
        if same:
            selected.append(sorted(same, key=lambda x: -x[0])[0][1])

        diff_candidates = []
        for regime, matches in by_regime.items():
            if regime != current_regime:
                diff_candidates.extend(matches)
        if diff_candidates:
            selected.append(sorted(diff_candidates, key=lambda x: -x[0])[0][1])

        selected_ids = {getattr(c, "case_id", id(c)) for c in selected}
        remaining = [c for s, c in sorted(all_matches, key=lambda x: -x[0])
                     if getattr(c, "case_id", id(c)) not in selected_ids]
        selected.extend(remaining[:max(0, top_k - len(selected))])

        return selected[:top_k]

    def stats(self) -> dict:
        return {
            "total_rules": len(self.rules),
            "rules_by_tier": self._tier_distribution(),
            "total_cases": len(self.cases),
            "gold_cases": len([c for c in self.cases if c.case_quality == "GOLD"]),
        }
